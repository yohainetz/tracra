from openpyxl import Workbook
from openpyxl.styles import Color, Fill, PatternFill
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import json
from tqdm import tqdm
from utils_resource import ResourceManager
import logging
import os

RESOURCE_CONFIG = ["MAIL_ID",
                   "LINK_COUNTER",
                   "DOMAIN_NUMBER",
                   "WITHOUT_SUBDOMAIN_NUMBER",
                   "HTML_TAG",
                   "HTML_TAG_ADDITIONAL",
                   "BLOCKED_BY",
                   "TRACKING_PIXEL",
                   "EMAIL_LEAKAGE",
                   "#HTTP_PARAMETER",
                   "PATH_LENGTH",
                   "#PATH_ELEMENTS",
                   "#SUBDOMAINS",
                   "URL_LENGTH",
                   "UNSUBSCRIBE_INDICATORS",
                   "DESCRIBE_PATH_ELES",
                   "DESCRIBE_PARAM_KEYS",
                   "DESCRIBE_PARAM_VALS",
                   "THIRD_PARTY",
                   "DFE_DOMAIN_CATEGORIZATION",
                   "DMOZ_DOMAIN_CATEGORIZATION",
                   "WHOIS_COUNTRY",
                   "WHOIS_OWNER",
                   "ALEXA_RANKING",
                   "TRANCO_RANKING",
                   "PRIMARY_ACTION",
                   "DISPLAYLINK_TEXT",
                   "DISPLAYLINK_#PARAMS",
                   "DISPLAYLINK_#SUBDOMAINS",
                   "DISPLAYLINK_URL_LENGTH",
                   "DISPLAYPATH_LENGTH",
                   "DISPLAY_#PATH_ELEMENTS",
                   "EDITDISTANCE_DISPLAYLINK_URL",
                   "EDITDISTANCE_DISPLAYLINKDOMAIN_URLDOMAIN",
                   "EDITDISTANCE_DISPLAYLINKDOMAIN_FROMDOMAIN",
                   "EDITDISTANCE_URLDOMAIN_FROMDOMAIN",
                   "TLD"  # TODO  privacy check
                   ]
RESOURCE_CONFIG_RAWDATA = [
                    "URL",
                    "PARENT_SUBJECT"
                    ]

RESOURCE_CONFIG_EXTRA = {
    "TLD": "yellow",
    "WHOIS_COUNTRY": "orange",
    "ALEXA_RANKING": "orange",
    "DMOZ_DOMAIN_CATEGORIZATION": "red",
    "DFE_DOMAIN_CATEGORIZATION": "red"
}

MAIL_CONFIG = ["FOLDER",
               "MAIL_ID",
               "SENDER_NUMBER",
               "SENDERDOMAIN_NUMBER",
               "WITHOUT_SUBDOMAIN_NUMBER",
               "#SENDER_SUBDOMAINS",
               "SENDER_LENGTH",
               "SENDERDOMAIN_LENGTH",
               "YEAR",
               "#TRACKING_URLS_IN_MAIL",
               "#NON_TRACKING_URLS_IN_MAIL",
               "MAILINGLIST_PARAMETER",
               "PERSONAL_SALUTATION",
               "SIGNATURE_BLOCK",
               "FROM_DISPLAY_LENGTH",
               "HTML_PLAIN",
               "LANGUAGE",
               "READABILITY",
               "#TEXT_CHARACTERS",
               "#TEXT_SPACES",
               "TEXT_DUPLICATE",
               "COMMON_MAIL_PROVIDER",
               "#MTAS",
               "#MTA_DIFFERENT_DOMAINS",
               "COMMON_EMAIL_MARKETING_SERVICE",
               "MAIN_ADDRESSEE",
               "TO",
               "FROM",
               "CC",
               "BCC",
               "FLAG_SEEN",
               "FLAG_ANSWERED",
               "FLAG_FLAGGED",
               "KEYWORDS",
               "RE_FWD",
               "#ATTACHMENTS",
               "SENDER_ALEXA_RANKING",
               "SENDER_TRANCO_RANKING",
               "DFE_DOMAIN_CATEGORIZATION",
               "DMOZ_DOMAIN_CATEGORIZATION",
               "WHOIS_COUNTRY",
               "WHOIS_OWNER",
               "RETURN_RECEIPT_HEADER",
               "SPF_DKIM_DMARC",
               "USERAGENT_XMAILER",
               "SENDER_TLD",  # TODO privacy check
               "MIME_STRUCTURE",
               "COUNT_HTML_TAGS"
               ]

MAIL_CONFIG_RAWDATA = [
                "RAW_FROM",
                "RAW_SUBJECT"
]

MAIL_CONFIG_EXTRA = {
    "USERAGENT_XMAILER": "yellow",
    "SENDER_TLD": "yellow",
    "WHOIS_COUNTRY": "orange",
    "SENDER_ALEXA_RANKING": "orange",
    "DMOZ_DOMAIN_CATEGORIZATION": "red",
    "DFE_DOMAIN_CATEGORIZATION": "red"
}

def build_style_list(header_list, extra_conf):
    sl = []

    colorcodes = {
        "yellow": "FEFF33",
        "orange": "FFA033",
        "red": "FF5733"
    }

    for key, val in extra_conf.items():
        col_num = header_list.index(key)
        sl.append((col_num, colorcodes[val]))
    return sl

def field_is_not_empty(field):
    return (field != "" or field.strip() != "") and  \
           (field is not None) and (field != "N/A")

def fill_empty_with_na(row):
    return list(map(lambda f: f if field_is_not_empty(f) else "N/A", row))

def writeMails(analyzed_mails_arr, destname, folderpath, rawdata=False):
    wb = Workbook()
    dest_filename = destname + '.xlsx'
    raw_dest_filename = "KLARTEXT_TRACKING_" + dest_filename
    ws_mails = wb.active
    ws_mails.title = "Mails #1"

    ws_resources = wb.create_sheet('Resources #1')
    print("Write to disk...")

    mail_count = 1
    res_count  = 1

    MC = MAIL_CONFIG
    RC = RESOURCE_CONFIG
    if rawdata:
        MC = MAIL_CONFIG_RAWDATA + MAIL_CONFIG
        RC = RESOURCE_CONFIG_RAWDATA + RESOURCE_CONFIG

    # WRITE sheet headers
    ws_mails.append(MC)
    ws_resources.append(RC)

    mail_style_list     = build_style_list(MC, MAIL_CONFIG_EXTRA)
    resource_style_list = build_style_list(RC, RESOURCE_CONFIG_EXTRA)

    for mail in tqdm(analyzed_mails_arr, unit=" E-Mails"):
        mail_row = list(map(mail.cache.get, MC))
        mail_row = fill_empty_with_na(mail_row)
        ws_mails.append(mail_row)
        mail_count += 1
        for msl in mail_style_list:
            try:
                if field_is_not_empty(mail_row[msl[0]]):
                    ws_mails.cell(row=mail_count, column=msl[0]+1).fill = PatternFill(fgColor=msl[1],
                                                                              fill_type='solid')
            except Exception as e:
                logging.debug(str(mail_row))
                logging.debug(e)

        for resource in mail.web_resources:
            try:
                resource_row = list(map(resource.cache.get, RC))
                resource_row = fill_empty_with_na(resource_row)
                ws_resources.append(resource_row)
                res_count += 1
                for sl in resource_style_list:
                    if field_is_not_empty(resource_row[sl[0]]):
                        ws_resources.cell(row=res_count, column=sl[0]+1).fill = PatternFill(fgColor=sl[1],
                                                                                    fill_type = 'solid')
            except Exception as e:
                logging.debug(str(resource_row))
                logging.debug(e)

    destpath = os.path.join(folderpath,dest_filename)
    wb.save(destpath)

def write_plaintext_tracking(data, destname, folderpath):
    sorted_data = sorted(data, key=lambda d: d[0].split("@")[-1])

    wb = Workbook()
    dest_filename = destname + '.xlsx'
    ws_tracking = wb.active
    ws_tracking.title = "Tracking Senders"

    ws_tracking.append(["Absendeadresse","Anzahl verdächtiger Emails"])

    for row in sorted_data:
        ws_tracking.append(row)

    wb.save(os.path.join(folderpath,dest_filename))
